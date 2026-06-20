"""
Serveur TCP Python ESP32 UWB
IEMN Stage 2026 le 27/03/2026

Objectif ;

Reçoit les mesures UWB en continu depuis l'ESP32 et les enregistre
automatiquement dans un fichier Excel .

"""

import socket
import datetime
import threading
import time

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# CONFIGURATION 
HOST        = '0.0.0.0'              # Écoute sur toutes les interfaces réseau
PORT        = 8080                   # Port TCP (identique dans le code Arduino)
EXCEL_FILE  = 'donnees_mesures.xlsx' # Chemin du fichier de sortie
SAVE_EVERY  = 10                     # Sauvegarde automatique toutes les N secondes

# Stockage en mémoire 
data_capteurs  = []   # Mesures brutes par capteur et par canal
data_positions = []   # Positions calculées par canal
data_moyennes  = []   # Positions finales moyennes
lock = threading.Lock()


# SAUVEGARDE EXCEL

def save_to_excel():
    """Enregistre toutes les données en mémoire dans un fichier Excel formaté."""
    with lock:
        if not data_capteurs and not data_positions and not data_moyennes:
            return
        # Copies locales pour libérer le verrou rapidement
        snap_c = list(data_capteurs)
        snap_p = list(data_positions)
        snap_m = list(data_moyennes)

    print(f"\n[SAVE] Sauvegarde dans '{EXCEL_FILE}' …")
    try:
        cols_c = ['N°', 'Timestamp (ms)', 'Heure', 'Canal', 'Freq (MHz)',
                  'Capteur ID', 'ToF (ps)', 'Distance (m)']
        cols_p = ['N°', 'Timestamp (ms)', 'Heure', 'Canal', 'Freq (MHz)',
                  'X (mm)', 'Y (mm)', 'X (m)', 'Y (m)', 'Distance Radiale (mm)']
        cols_m = ['N°', 'Timestamp (ms)', 'Heure',
                  'X (mm)', 'Y (mm)', 'X (m)', 'Y (m)', 'Distance Radiale (mm)']

        def make_df(rows, cols):
            df = pd.DataFrame(rows, columns=cols[1:])
            df.insert(0, 'N°', range(1, len(df) + 1))
            return df

        df_c = make_df(snap_c, cols_c)
        df_p = make_df(snap_p, cols_p)
        df_m = make_df(snap_m, cols_m)

        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
            df_c.to_excel(writer, sheet_name='Mesures_Capteurs',   index=False)
            df_p.to_excel(writer, sheet_name='Positions_Canaux',   index=False)
            df_m.to_excel(writer, sheet_name='Positions_Moyennes', index=False)
            for name in ('Mesures_Capteurs', 'Positions_Canaux', 'Positions_Moyennes'):
                _style_sheet(writer.book[name])

        print(f"[SAVE] OK — {len(snap_c)} mesures · "
              f"{len(snap_p)} positions · {len(snap_m)} moyennes")

    except PermissionError:
        print("[SAVE] ⚠  Fermez le fichier Excel avant la prochaine sauvegarde !")
    except Exception as e:
        print(f"[SAVE] Erreur : {e}")


def _style_sheet(ws):
    """Applique un en-tête coloré (bleu IEMN) et ajuste les largeurs de colonnes."""
    fill = PatternFill(start_color="003E7E", end_color="003E7E", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        max_w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 3, 28)


def _auto_save_loop():
    """Thread de sauvegarde périodique (toutes les SAVE_EVERY secondes)."""
    while True:
        time.sleep(SAVE_EVERY)
        save_to_excel()



def _parse_line(line: str, ts_local: str) -> bool:
    """
    Analyse une ligne CSV reçue de l'ESP32 et ajoute les données dans
    la liste correspondante.
    Retourne True si la ligne a été reconnue et traitée.
    """
    parts = [p.strip() for p in line.split(',')]

    # Ligne POSITION MOYENNE : millis,MOYENNE,,MOY,x_mm,y_mm,radial_mm 
    if len(parts) >= 7 and parts[1] == 'MOYENNE' and parts[3] == 'MOY':
        try:
            ts, x_mm, y_mm, radial = parts[0], float(parts[4]), float(parts[5]), float(parts[6])
            with lock:
                data_moyennes.append([ts, ts_local,
                                       x_mm, y_mm,
                                       round(x_mm / 1000.0, 3),
                                       round(y_mm / 1000.0, 3),
                                       radial])
            print(f"  [MOY] ({x_mm:.1f} mm, {y_mm:.1f} mm) | Radiale: {radial:.1f} mm")
            return True
        except ValueError:
            return False

    # Ligne POSITION par canal : millis,canal,freq,POS,x_mm,y_mm,radial_mm
    if len(parts) >= 7 and parts[3] == 'POS':
        try:
            ts, canal, freq = parts[0], parts[1], parts[2]
            x_mm, y_mm, radial = float(parts[4]), float(parts[5]), float(parts[6])
            with lock:
                data_positions.append([ts, ts_local, canal, freq,
                                        x_mm, y_mm,
                                        round(x_mm / 1000.0, 3),
                                        round(y_mm / 1000.0, 3),
                                        radial])
            print(f"  [POS] Canal {canal} ({freq} MHz) → ({x_mm:.1f}, {y_mm:.1f}) mm")
            return True
        except ValueError:
            return False

    # Ligne CAPTEUR
    if len(parts) >= 6:
        try:
            ts, canal, freq = parts[0], parts[1], parts[2]
            capteur         = parts[3].upper()
            tof             = float(parts[4])
            distance        = float(parts[5])
            with lock:
                data_capteurs.append([ts, ts_local, canal, freq,
                                       capteur, tof, distance])
            print(f"  [CAP] {capteur}  Canal {canal} ({freq} MHz) → {distance:.3f} m  ToF: {tof:.2f} ps")
            return True
        except ValueError:
            pass

    return False


# GESTION DES CONNEXIONS 

def _handle_client(conn: socket.socket, addr):
    """Gère la session TCP d'un ESP32 connecté."""
    print(f"\n[+] ESP32 connecté depuis {addr[0]}:{addr[1]}")
    buffer = ""
    try:
        while True:
            chunk = conn.recv(4096)
            if not chunk:
                break
            # Accumulation dans un buffer — on découpe uniquement aux '\n'
            buffer += chunk.decode('utf-8', errors='replace')
            while '\n' in buffer:
                line, buffer = buffer.split('\n', 1)
                line = line.strip()
                if line:
                    ts_local = datetime.datetime.now().strftime("%H:%M:%S.%f")[:-3]
                    _parse_line(line, ts_local)
    except ConnectionResetError:
        print(f"[-] ESP32 {addr[0]} déconnecté brutalement.")
    except Exception as e:
        print(f"[ERREUR client] {e}")
    finally:
        conn.close()
        print(f"[-] Connexion fermée avec {addr[0]} — sauvegarde…")
        save_to_excel()


# POINT D'ENTRÉE 

def start_server():
    """Lance le serveur TCP, accepte les connexions des ESP32 en parallèle."""
    server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    server.bind((HOST, PORT))
    server.listen(5)

    # Thread de sauvegarde automatique (daemon → s'arrête avec le programme)
    threading.Thread(target=_auto_save_loop, daemon=True).start()

    print("=" * 60)
    print("  Serveur UWB → Excel  (IEMN Stage 2026)")
    print(f"  Port TCP       : {PORT}")
    print(f"  Fichier Excel  : {EXCEL_FILE}")
    print(f"  Sauvegarde auto: toutes les {SAVE_EVERY} secondes")
    print("  Appuyez sur Ctrl+C pour arrêter et sauvegarder.")
    print("=" * 60)

    try:
        while True:
            conn, addr = server.accept()
            threading.Thread(target=_handle_client,
                             args=(conn, addr), daemon=True).start()
    except KeyboardInterrupt:
        print("\n[INFO] Arrêt demandé — sauvegarde finale…")
        save_to_excel()
    finally:
        server.close()
        print("[INFO] Serveur arrêté.")


if __name__ == "__main__":
    start_server()

